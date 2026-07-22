"""
serp_scraper.py

Renamed from top3.py (moved into scripts/ so every pipeline stage lives
together) -- fetches the top-3 Google organic results (url + title) for
each keyword in an input sheet, using a real Chrome browser with several
tabs open at once, round-robin.

Content/behavior is otherwise UNCHANGED from top3.py, including the
column-matching (`_find_columns` looks for a "Keywords" header) and the
default EXCEL_FILE path -- those are kept exactly as last edited.

The ONE change: run_search_pool() now accepts an optional `on_result`
callback, called the instant each tab's job finishes (keyword, results,
start_time, stop_time), BEFORE it's written to `output_path` (which is
now also optional). This is what lets scripts/run_pipeline.py reuse this
EXACT function -- tabs, retries, timeouts, everything -- by passing a
callback in, instead of a second copy of the tab-management loop having
to be written and kept in sync by hand.

Run standalone (writes straight to `<EXCEL_FILE>_top3.csv`, same as
before):
    python -m scripts.serp_scraper
"""

import csv
import json
import os
import time
import random
import zipfile
import tempfile
from urllib.parse import quote_plus
from datetime import datetime
from typing import Optional
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, StaleElementReferenceException
)
from selenium.webdriver.chrome.service import Service

EXCEL_FILE = "backend/datasets/"
DELAY_MIN     = 5
DELAY_MAX     = 12
NUM_TABS      = 10
TAB_TIMEOUT   = 50


def create_proxy_auth_extension(proxy_url: str) -> str:
    """Helper to dynamically generate a Chrome extension zip that handles proxy authentication."""
    from urllib.parse import urlparse
    parsed = urlparse(proxy_url)
    host = parsed.hostname
    port = parsed.port or 80
    username = parsed.username
    password = parsed.password
    
    manifest_json = """
    {
        "version": "1.0.0",
        "manifest_version": 2,
        "name": "Chrome Proxy",
        "permissions": [
            "proxy",
            "tabs",
            "unlimitedStorage",
            "storage",
            "<all_urls>",
            "webRequest",
            "webRequestBlocking"
        ],
        "background": {
            "scripts": ["background.js"]
        },
        "minimum_chrome_version":"22.0.0"
    }
    """

    background_js = f"""
    var config = {{
        mode: "fixed_servers",
        rules: {{
          singleProxy: {{
            scheme: "http",
            host: "{host}",
            port: parseInt({port})
          }},
          bypassList: []
        }}
      }};

    chrome.proxy.settings.set({{value: config, scope: "regular"}}, function() {{}});

    chrome.webRequest.onAuthRequired.addListener(
        function callback(details) {{
            return {{
                authCredentials: {{
                    username: "{username}",
                    password: "{password}"
                }}
            }};
        }},
        {{urls: ["<all_urls>"]}},
        ['blocking']
    );
    """
    
    temp_dir = tempfile.gettempdir()
    plugin_file = os.path.join(temp_dir, f"proxy_auth_plugin_{port}.zip")
    
    with zipfile.ZipFile(plugin_file, 'w') as zp:
        zp.writestr("manifest.json", manifest_json)
        zp.writestr("background.js", background_js)
        
    return plugin_file


_KEYWORD_COLUMN_NAMES = ("keywords", "keyword", "kw")


def _find_columns(headers: list[str]):
    keyword_col = None
    for i, h in enumerate(headers):
        if h.lower() in _KEYWORD_COLUMN_NAMES:
            keyword_col = i

    if keyword_col is None:
        raise ValueError(f"'Keyword' column not found. Headers: {headers}")

    return keyword_col


def load_excel(path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    print(f"  Headers found: {headers}\n")

    keyword_col = _find_columns(headers)

    print(f"  Keyword col : col {keyword_col + 1} → '{headers[keyword_col]}'")

    rows = []
    for row in ws.iter_rows(min_row=2):
        kw = str(row[keyword_col].value).strip() if row[keyword_col].value else ""
        if kw:
            rows.append({"keyword": kw})

    print(f"\n  Total rows to process: {len(rows)}\n")
    return rows


def load_csv(path: str) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        all_rows = list(reader)

    if not all_rows:
        raise ValueError(f"CSV file is empty: {path}")

    headers = [h.strip() for h in all_rows[0]]
    data_rows = all_rows[1:]
    print(f"  Headers found: {headers}\n")

    keyword_col = _find_columns(headers)

    print(f"  Keyword col : col {keyword_col + 1} → '{headers[keyword_col]}'")

    rows = []
    for line in data_rows:
        kw = line[keyword_col].strip() if len(line) > keyword_col and line[keyword_col] else ""
        if kw:
            rows.append({"keyword": kw})

    print(f"\n  Total rows to process: {len(rows)}\n")
    return rows


def load_data(path: str) -> list[dict]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return load_csv(path)
    elif ext in (".xlsx", ".xlsm"):
        return load_excel(path)
    else:
        raise ValueError(f"Unsupported file type '{ext}'. Use .csv or .xlsx.")


def get_driver() -> webdriver.Chrome:
    options = Options()
    
    # Configure proxy if SCRAPING_PROXY is defined
    proxy_url = os.environ.get("SCRAPING_PROXY")
    if proxy_url:
        try:
            plugin_file = create_proxy_auth_extension(proxy_url)
            options.add_extension(plugin_file)
            print(f"[serp_scraper] Configured proxy extension via {proxy_url.split('@')[-1]}")
        except Exception as e:
            print(f"[serp_scraper] Warning: failed to configure proxy extension: {e}")

    options.page_load_strategy = "eager"
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        service = Service(ChromeDriverManager().install())
    except ImportError:
        service = Service()

    driver = webdriver.Chrome(service=service, options=options)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


def accept_consent(driver):
    try:
        btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//button[.//span[contains(text(),'Accept all')]]")
            )
        )
        btn.click()
        time.sleep(1)
    except TimeoutException:
        pass


def extract_results(driver) -> list[dict]:
    anchors = driver.find_elements(By.CSS_SELECTOR, 'a[jsname="UWckNb"]')
    if not anchors:
        anchors = driver.find_elements(By.CSS_SELECTOR, 'div.g a[href]')

    results = []
    for a in anchors:
        try:
            href = a.get_attribute("href")
            if not href:
                continue
            if any(s in href for s in [
                "webcache", "google.com/search", "google.com/url",
                "accounts.google", "javascript:", "#:~:text="
            ]):
                continue
            if not href.startswith("http"):
                continue

            try:
                title = a.find_element(By.TAG_NAME, "h3").text.strip()
            except NoSuchElementException:
                title = ""
            if not title:
                title = (a.get_attribute("aria-label") or "").strip()
            if not title:
                title = a.text.strip()
            if not title:
                continue   # not a real organic result (icon/empty link etc.)

            results.append({"url": href, "title": title})
        except StaleElementReferenceException:
            continue
    return results


def start_search(driver, keyword: str):
    driver.get(f"https://www.google.com/search?q={quote_plus(keyword)}&num=10&hl=en")


def tab_ready(driver) -> bool:
    """Non-blocking check: has this tab's results page rendered yet?
    Used instead of WebDriverWait so checking one tab never blocks the others."""
    return bool(driver.find_elements(By.ID, "search"))


def open_tabs(driver, count: int) -> list[str]:
    handles = driver.window_handles
    while len(handles) < count:
        driver.switch_to.window(handles[-1])
        driver.execute_script("window.open('about:blank', '_blank');")
        handles = driver.window_handles
    return handles[:count]


class TabJob:
    def __init__(self, row: dict):
        self.row        = row
        self.start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.deadline   = time.time() + TAB_TIMEOUT
        self.retried    = False


def run_search_pool(driver, rows: list, output_path: Optional[str] = None, on_result=None, num_tabs: int = NUM_TABS):
    """Tab round-robin loop -- IDENTICAL logic to the original top3.py.

    `output_path`, if given, gets the same 4-column CSV
    (Keyword, Start Time, Stop Time, Top 3 URLs (JSON)) written
    incrementally, exactly as before.

    `on_result`, if given, is called as on_result(row, results,
    start_time, stop_time) the instant each job finishes -- BEFORE the
    CSV write -- so a caller (e.g. run_pipeline.py) can hand the result
    off to its own downstream processing (category/landing-blog/info-
    comm) without needing a second copy of this loop. `row` is the full
    dict from `rows` (not just the keyword string), so callers can stash
    extra fields (like a pipeline row id) on each row before calling this.
    """
    handles = open_tabs(driver, num_tabs)
    print(f"  Opened {len(handles)} browser tabs\n")

    # Warm up tab 1 so the "Accept all" consent screen is dismissed once,
    # instead of every tab hitting it independently.
    driver.switch_to.window(handles[0])
    driver.get("https://www.google.com")
    accept_consent(driver)

    queue: list = list(rows)
    total       = len(rows)
    done        = 0
    jobs: dict[str, Optional[TabJob]] = {h: None for h in handles}
    cooldown: dict[str, float] = {h: 0.0 for h in handles}

    def label(h):
        return f"[Tab {handles.index(h) + 1}]"

    out_file = open(output_path, "w", newline="", encoding="utf-8-sig") if output_path else None
    writer = None
    try:
        if out_file:
            writer = csv.writer(out_file)
            writer.writerow(["Keyword", "Start Time", "Stop Time", "Top 3 URLs (JSON)"])

        while queue or any(jobs[h] is not None for h in handles):
            for h in handles:
                driver.switch_to.window(h)
                job = jobs[h]

                if job is None:
                    if not queue or time.time() < cooldown[h]:
                        continue

                    row = queue.pop(0)
                    print(f"{label(h)} [{done + 1}/{total}] \"{row['keyword']}\"")
                    start_search(driver, row["keyword"])
                    jobs[h] = TabJob(row)
                    continue

                if not tab_ready(driver):
                    if time.time() > job.deadline and not job.retried:
                        accept_consent(driver)   # results may be stuck behind a consent screen
                        job.retried = True
                        job.deadline = time.time() + TAB_TIMEOUT
                    elif time.time() > job.deadline:
                        print(f"{label(h)}            Timed out waiting for results. Triggering Firecrawl search fallback...")
                        stop_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        
                        results = []
                        try:
                            from scripts.firecrawl_scraper import fetch_top_results_via_firecrawl
                            fc_data = fetch_top_results_via_firecrawl(job.row["keyword"])
                            results = [{"url": r["url"], "title": r["title"]} for r in fc_data.get("top_3", [])]
                        except Exception as fe:
                            print(f"{label(h)}            Firecrawl fallback failed: {fe}")
                            
                        if on_result:
                            on_result(job.row, results, job.start_time, stop_time)
                        if writer:
                            writer.writerow([job.row["keyword"], job.start_time, stop_time, json.dumps(results, ensure_ascii=False)])
                            out_file.flush()
                        done += 1
                        cooldown[h] = time.time() + random.uniform(DELAY_MIN, DELAY_MAX)
                        jobs[h] = None
                    continue

                results = extract_results(driver)[:5]
                if not results:
                    print(f"{label(h)}            No results found. Triggering Firecrawl search fallback...")
                    try:
                        from scripts.firecrawl_scraper import fetch_top_results_via_firecrawl
                        fc_data = fetch_top_results_via_firecrawl(job.row["keyword"])
                        results = [{"url": r["url"], "title": r["title"]} for r in fc_data.get("top_3", [])]
                    except Exception as fe:
                        print(f"{label(h)}            Firecrawl fallback failed: {fe}")

                print(f"{label(h)}            Top 3 URLs:")
                for idx, r in enumerate(results[:3], 1):
                    print(f"{label(h)}              {idx}. {r['title']} — {r['url']}")

                stop_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if on_result:
                    on_result(job.row, results, job.start_time, stop_time)
                if writer:
                    writer.writerow([job.row["keyword"], job.start_time, stop_time, json.dumps(results, ensure_ascii=False)])
                    out_file.flush()

                done += 1
                print(f"{label(h)} [{done}/{total}] Done\n")

                cooldown[h] = time.time() + random.uniform(DELAY_MIN, DELAY_MAX)
                jobs[h] = None

            time.sleep(0.5)
    finally:
        if out_file:
            out_file.close()


# ─────────────────────────────────────────────
# MAIN (standalone run -- SERP fetch only, no category/landing-blog/info-comm)
# ─────────────────────────────────────────────
def main():
    output_path = os.path.splitext(EXCEL_FILE)[0] + "_top3.csv"

    print(f"\n{'='*60}")
    print(f"  Google Top-3 URL Fetcher")
    print(f"  File    : {EXCEL_FILE}")
    print(f"  Output  : {output_path}")
    print(f"  Tabs    : {NUM_TABS} tabs in one browser, searched round-robin")
    print(f"  Started : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")

    rows = load_data(EXCEL_FILE)

    if not rows:
        print("No rows found. Exiting.")
        return

    driver = get_driver()
    try:
        run_search_pool(driver, rows, output_path)
    finally:
        driver.quit()

    print(f"\n{'='*60}")
    print(f"  DONE — {len(rows)} keywords processed")
    print(f"  Results : {output_path}")
    print(f"  Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
